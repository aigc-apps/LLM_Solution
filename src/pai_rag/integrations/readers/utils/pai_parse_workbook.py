from typing import List
import numpy as np
import openpyxl
import pandas as pd
from llama_index.core.schema import Document

from loguru import logger

import re

IMAGE_REGEX = r'<img src="([^"]+)" alt="PAIRAG_IMAGE_TAG">\n'


class PaiFormData:
    def __init__(self, title, data, header_row: int = -1):
        self.title = title
        self.data = data
        self.header_row = header_row


def split_row_group(row_group, headers=[], splitter=None, form_title=None):
    """
    Split a row group into smaller pieces.
    """
    raw_text = ""
    form_title = form_title + "\n\n"
    title_text = ""
    if len(row_group) == 0:
        return []

    if len(headers) > 0:
        assert len(headers) == len(
            row_group[0]
        ), f"Header and row data length mismatch! headers: {headers}, row: {row_group[0]}"

    is_outline_column = []
    for j in range(len(row_group[0])):
        first_value = row_group[0][j]
        if not first_value:
            is_outline_column.append(False)
            continue

        is_same_value = True
        for i in range(1, len(row_group)):
            if row_group[i][j] != first_value:
                is_same_value = False
                break

        if is_same_value:
            if len(headers) == 0:
                column_text = f"{first_value}\n\n\n"
            else:
                column_text = f"{headers[j]}: {first_value}\n\n\n"

            if len(column_text) <= 30:
                title_text += column_text
            else:
                is_same_value = False

        is_outline_column.append(is_same_value)

    for i in range(len(row_group)):
        for j in range(len(row_group[0])):
            if is_outline_column[j]:
                continue
            else:
                if len(headers) == 0 or headers[j] is None:
                    if not row_group[i][j]:
                        continue
                    else:
                        raw_text += f"{row_group[i][j]}\n"
                else:
                    raw_text += f"{headers[j]}: {row_group[i][j]}\n"

        raw_text += "\n\n"

    image_url_list = re.findall(IMAGE_REGEX, raw_text) + re.findall(
        IMAGE_REGEX, title_text
    )

    image_info_list = []
    for image_url in image_url_list:
        image_info_list.append({"image_url": image_url})

    raw_text = re.sub(IMAGE_REGEX, "", raw_text)
    title_text = re.sub(IMAGE_REGEX, "", title_text)

    if len(raw_text) < 3000:
        return [
            Document(
                text=form_title + title_text + raw_text,
                extra_info={"image_info_list": image_info_list},
            )
        ]
    else:
        return [
            Document(
                text=form_title + title_text + split,
                extra_info={"image_info_list": image_info_list},
            )
            for split in splitter.split_text(raw_text)
        ]


def chunk_form(form_title, form_data, header_row=-1, splitter=None):
    """
    Chunk the form data into smaller pieces.
    """
    chunks = []

    # Clean data
    columns = None
    if header_row != -1:
        columns = form_data[header_row]
        form_data = form_data[header_row + 1 :]

    df = pd.DataFrame(form_data, columns=columns)
    # 去重
    df = df.drop_duplicates()
    df = df.T.drop_duplicates().T

    # 去掉重复的na
    df = df.dropna(how="all")
    df = df.dropna(how="all", axis=1)

    if df.empty:
        return chunks

    df = df.fillna("")
    if columns is not None and len(columns) > 0:
        columns = df.columns
    else:
        columns = []
    values = df.to_numpy()
    i = 0
    while i < values.shape[0]:
        row_group_values = []
        current_size = 0
        for j in range(values.shape[1]):
            if values[i][j] is not None:
                current_size += len(str(values[i][j]))
        row_group_values.append(values[i])

        # 试探下一行是否可以合并
        while i + 1 < values.shape[0]:
            should_merge = False
            if values.shape[1] <= 1:
                should_merge = True
            else:
                for j in range(values.shape[1]):
                    if (
                        values[i + 1][j] is not None
                        and values[i + 1][j] != ""
                        and values[i + 1][j] == values[i][j]
                    ):
                        should_merge = True
                        break

            if should_merge:
                row_group_values.append(values[i + 1])
                i += 1
                continue
            else:
                chunks.extend(
                    split_row_group(
                        row_group_values, columns, splitter, form_title=form_title
                    )
                )
                row_group_values = []
                break

        if len(row_group_values) > 0:
            chunks.extend(
                split_row_group(
                    row_group_values, columns, splitter, form_title=form_title
                )
            )
        i += 1

    return chunks


def is_empty(value):
    if value is None:
        return True

    if isinstance(value, str) and value.strip() == "":
        return True

    return False


def dfs(data, i, j, visited):
    ltop_x, ltop_y, rbottom_x, rbottom_y = i, j, i, j

    st = [(i, j)]
    while len(st) > 0:
        i, j = st.pop()

        visited[i][j] = 1

        if i >= 1 and not is_empty(data[i - 1][j]) and not visited[i - 1][j]:
            ltop_x = min(ltop_x, i - 1)
            st.append((i - 1, j))
        elif i >= 1:
            visited[i - 1][j] = 1

        if j >= 1 and not is_empty(data[i][j - 1]) and not visited[i][j - 1]:
            ltop_y = min(ltop_y, j - 1)
            st.append((i, j - 1))
        elif j >= 1:
            visited[i][j - 1] = 1

        if (
            j < data.shape[1] - 1
            and not is_empty(data[i][j + 1])
            and not visited[i][j + 1]
        ):
            rbottom_y = max(rbottom_y, j + 1)
            st.append((i, j + 1))
        elif j < data.shape[1] - 1:
            visited[i][j + 1] = 1

        if (
            i < data.shape[0] - 1
            and not is_empty(data[i + 1][j])
            and not visited[i + 1][j]
        ):
            rbottom_x = max(rbottom_x, i + 1)
            st.append((i + 1, j))
        elif i < data.shape[0] - 1:
            visited[i + 1][j] = 1

        if len(st) == 0:
            for i in range(ltop_x, rbottom_x):
                for j in range(ltop_y, rbottom_y):
                    if not is_empty(data[i][j]) and visited[i][j] == 0:
                        st.append((i, j))

    return ltop_x, ltop_y, rbottom_x, rbottom_y


def vsplit(data, x1, y1, x2, y2, merge_state):
    forms = []
    x = x1
    start_x = x1

    while x <= x2:
        header = ""
        head_value = find_super_header(data, x, y1, y2, merge_state)
        while x <= x2 and head_value is not None:
            header += f"{head_value}\n"
            x += 1
            if x <= x2:
                head_value = find_super_header(data, x, y1, y2, merge_state)

        if x > x2:
            break

        start_x = x
        header_row = x - start_x
        header_length_total = 0
        header_length_count = 0
        for y in range(y1, y2 + 1):
            if isinstance(data[x][y], str):
                header_length_total += len(
                    data[x][y].replace(" ", "").replace("\n", "").strip()
                )
                header_length_count += 1
            else:
                header_length_total += 30
                header_length_count += 1

        if header_length_count > 0 and header_length_total / header_length_count >= 10:
            header_row = -1

        while x <= x2 and find_super_header(data, x, y1, y2, merge_state) is None:
            x += 1

        end_x = x
        # print(start_x, end_x, y1, y2+1, header_row, header_length_total, header_length_count)

        forms.append(
            PaiFormData(
                title=header,
                data=data[start_x:end_x, y1 : y2 + 1],
                header_row=header_row,
            )
        )

        start_x = end_x
        x = start_x
    return forms


def hsplit(data, x1, y1, x2, y2, merge_state):
    forms = []
    start_y = y1
    cur_y = start_y + 1
    while cur_y < y2:
        max_table_length = -1
        continuous_none_count = np.zeros(x2 - x1 + 1)
        last_value_index = x1
        for x in range(x1, x2 + 1):
            if data[x][cur_y] is None:
                if x > x1:
                    continuous_none_count[x - x1] = (
                        continuous_none_count[x - x1 - 1] + 1
                    )
                else:
                    continuous_none_count[x - x1] = 1
            else:
                continuous_none_count[x - x1] = 0
                max_continuous_none_count = max(continuous_none_count)
                last_value_index = x
        max_table_length = max(max_table_length, last_value_index)
        if cur_y < y2 and (
            max_continuous_none_count > (last_value_index - x1 + 1) * 0.75
            or last_value_index <= max_table_length * 0.2
            or last_value_index <= 1
        ):
            if start_y < cur_y:
                logger.info(
                    f"Excel vsplit box: ({x1},{start_y}) => ({x2}, {y2}).",
                    x1,
                    start_y,
                    x2,
                    cur_y - 1,
                )
                forms.extend(vsplit(data, x1, start_y, x2, cur_y - 1, merge_state))
                start_y = cur_y + 1
                cur_y = start_y

        cur_y = cur_y + 1

    if start_y <= y2:
        logger.info(
            f"Excel vsplit box: ({x1},{start_y}) => ({x2}, {y2}).", x1, start_y, x2, y2
        )
        forms.extend(vsplit(data, x1, start_y, x2, y2, merge_state))

    return forms


def find_super_header(data, x, y1, y2, merge_state):
    if y2 - y1 < 1:
        return None

    unique_value = None
    for y in range(y1, y2 + 1):
        if data[x][y] is not None:
            if merge_state[x][y] == 0 or (unique_value and unique_value != data[x][y]):
                return None
            elif unique_value is None:
                unique_value = data[x][y]
    return unique_value


def split_sheet_v2(sheet, oss_client, splitter):
    """
    Split a sheet into a list of form object.
    """
    sheet_data = []
    for row in sheet.values:
        sheet_data.append(list(row))

    if len(sheet._images) > 0 and oss_client is None:
        logger.warning("No OSS provider registered, image will not be uploaded to OSS.")
    elif len(sheet._images) > 0:
        for image in sheet._images:
            anchor = image.anchor
            i = anchor._from.row
            j = anchor._from.col

            image_url = None
            try:
                image_url = oss_client.put_object_if_not_exists(
                    data=image._data(),
                    file_ext=f".{image.format}",
                    headers={
                        "x-oss-object-acl": "public-read"
                    },  # set public read to make image accessible
                    path_prefix="pairag/images/",
                )
                logger.info(f"Uploaded image to {image_url}.")
            except Exception as ex:
                logger.warning(f"Error occurred when upload image to OSS: {ex}")
                continue

            if len(sheet_data) == 0:
                sheet_data = [[]]
            if len(sheet_data[0]) == 0:
                sheet_data[0].append(None)

            if i >= len(sheet_data):
                i = len(sheet_data) - 1
            if j >= len(sheet_data[0]):
                j = len(sheet_data[0]) - 1
            if sheet_data[i][j] is None or not isinstance(sheet_data[i][j], str):
                sheet_data[i][
                    j
                ] = f"""<img src="{image_url}" alt="PAIRAG_IMAGE_TAG">\n"""
            else:
                sheet_data[i][
                    j
                ] += f"""<img src="{image_url}" alt="PAIRAG_IMAGE_TAG">\n"""

    data = np.array(sheet_data)
    merge_state = np.zeros_like(sheet_data)
    merged_ranges = sheet.merged_cells.ranges
    merge_bounds = []
    for merge in merged_ranges:
        merge_bounds.append(merge.bounds)
    merge_bounds = sorted(merge_bounds)
    for bound in merge_bounds:
        first_value = data[bound[1] - 1, bound[0] - 1]
        data[bound[1] - 1 : bound[3], bound[0] - 1 : bound[2]] = first_value
        merge_state[bound[1] - 1 : bound[3], bound[0] - 1 : bound[2]] = 1

    forms: List[PaiFormData] = []
    visited = np.zeros_like(data)
    for i in range(data.shape[0]):
        for j in range(data.shape[1]):
            if not visited[i][j] and data[i][j] is not None:
                ltop_x, ltop_y, rbottom_x, rbottom_y = dfs(data, i, j, visited)
                logger.info(
                    f"Excel hsplit box: ({ltop_x},{ltop_y}) => ({rbottom_x}, {rbottom_y})."
                )
                forms.extend(
                    hsplit(data, ltop_x, ltop_y, rbottom_x, rbottom_y, merge_state)
                )

            visited[i][j] = 1

    docs = []
    for form in forms:
        for doc in chunk_form(
            form.title, form.data, form.header_row, splitter=splitter
        ):
            doc.extra_info["sheet_name"] = sheet.title
            docs.append(doc)

    return docs


def parse_workbook(workbook_file, oss_client, splitter):
    docs = []
    workbook = openpyxl.open(workbook_file, data_only=True)
    sheetnames = workbook.sheetnames
    for sheetname in sheetnames:
        logger.info(f"Parsing sheet {sheetname}.")
        sheet = workbook[sheetname]
        splitted_docs = split_sheet_v2(sheet, oss_client=oss_client, splitter=splitter)
        docs.extend(splitted_docs)
        logger.info(f"Parsed sheet {sheetname}.")

    return docs
